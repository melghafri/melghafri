from netmiko import ConnectHandler
import openpyxl
import datetime

open_t = open('asw.txt')
read_t = open_t.read()
switch_ips = read_t.splitlines()



worksheet = openpyxl.load_workbook('Switch_Inlet_Temp.xlsx')
sheet_obj = worksheet.active
sheet_obj.cell(1, 1).value = 'Switch Name'
sheet_obj.cell(1, 2).value = datetime.date.today()

k = 2

for x in switch_ips:
    try:
        switch_connection_parameters = {
            "ip": x,
            "device_type": "cisco_ios",
            "username": "", # username
            "password": "" # password
        }
        ssh_session = ConnectHandler(**switch_connection_parameters)
        hostname = (ssh_session.find_prompt()).strip(">")
        show_inv = ssh_session.send_command("show inventory", use_textfsm=True)
        switch_model = show_inv[0]['pid']
        print(hostname)
        if "C9300" in switch_model:
            show_env_temp = ssh_session.send_command("show environment temperature", use_textfsm=True)
            show_sw_det = ssh_session.send_command("show switch detail", use_textfsm=True)
            stack_count = len(show_sw_det)
            for r in range(1, stack_count):
                sheet_obj.cell(k, 1).value = hostname
                sheet_obj.cell(k, 2).value = int(show_env_temp[r]['inlet_temperature_value'])
                k = k + 1
    except:
        sheet_obj.cell(k, 1).value = x
        sheet_obj.cell(k, 2).value = "switch is unreachable"
        k = k + 1

    worksheet.save("Switch_Inlet_Temp.xlsx")

