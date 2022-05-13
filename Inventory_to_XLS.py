import openpyxl
from netmiko import ConnectHandler

open_t = open('asw.txt')
read_t = open_t.read()
switch_ips = read_t.splitlines()

worksheet = openpyxl.load_workbook("Inventory.xlsx")
sheet_obj = worksheet.active
sheet_obj.cell(1, 1).value = 'HOSTNAME'
sheet_obj.cell(1, 2).value = 'name'
sheet_obj.cell(1, 3).value = 'descr'
sheet_obj.cell(1, 4).value = 'pid'
sheet_obj.cell(1, 5).value = 'vid'
sheet_obj.cell(1, 6).value = 'sn'

k = 2

for x in switch_ips:
    switch_connection_parameters = {
        "ip": x,
        "device_type": "cisco_ios",
        "username": "", # username
        "password": "" # password
    }

    ssh_session = ConnectHandler(**switch_connection_parameters)
    hostname = (ssh_session.find_prompt()).strip(">")
    print(f"======= CONNECTED TO {hostname} =======")
    show_inv = ssh_session.send_command("show inventory", use_textfsm=True)

    for y in range(len(show_inv)):
        sheet_obj.cell(k, 1).value = hostname
        sheet_obj.cell(k, 2).value = show_inv[y]['name']
        sheet_obj.cell(k, 3).value = show_inv[y]['descr']
        sheet_obj.cell(k, 4).value = show_inv[y]['pid']
        sheet_obj.cell(k, 5).value = show_inv[y]['vid']
        sheet_obj.cell(k, 6).value = show_inv[y]['sn']
        k = k + 1

    worksheet.save("Inventory.xlsx")
