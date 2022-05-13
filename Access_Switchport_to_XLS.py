import openpyxl
import re
from netmiko import ConnectHandler
from mac_vendor_lookup import MacLookup
from openpyxl.styles import Alignment

open_t = open('asw.txt')
read_t = open_t.read()
switch_ips = read_t.splitlines()


worksheet = openpyxl.load_workbook('scan.xlsx')
sheet_obj = worksheet.active
sheet_obj.cell(1, 1).value = 'Switch Name'
sheet_obj.cell(1, 2).value = 'Port'
sheet_obj.cell(1, 3).value = 'Description'
sheet_obj.cell(1, 4).value = 'Status'
sheet_obj.cell(1, 5).value = 'Mode'
sheet_obj.cell(1, 6).value = 'Voice VLAN'
sheet_obj.cell(1, 7).value = 'Data VLAN'
sheet_obj.cell(1, 8).value = 'Endpoint(s) MAC'
sheet_obj.cell(1, 9).value = 'Endpoint(s) MAC Lookup'


def find_switch_port_des(f_show_int_des, f_switch_port):
    for p in range(len(f_show_int_des)):
        show_des_port = f_show_int_des[p]['port']
        if f_switch_port == show_des_port:
            return f_show_int_des[p]['descrip']


def find_switch_port_trunk_vlans(f_show_int_switchport, f_switch_port):
    for p in range(len(f_show_int_switchport)):
        show_int_switchport_int = f_show_int_switchport[p]['interface']
        if f_switch_port == show_int_switchport_int:
            vlan_list_one_string = f_show_int_switchport[p]['trunking_vlans']
            vlan_string = vlan_list_one_string[0]
            vlan_list = vlan_string.split(',')
            return vlan_list


def expand_trunk_vlan_list(f_trunk_vlan_list):
    expanded_list = []
    for z in f_trunk_vlan_list:
        if '-' in z:
            unpack = z.split('-')
            start = int(unpack[0])
            finish = int(unpack[1]) + 1
            for b in range(start, finish):
                expanded_list.append(int(b))
        else:
            expanded_list.append(int(z))
    return expanded_list


def find_switchport_mode(f_show_int_switchport, f_switch_port):
    for p in range(len(f_show_int_switchport)):
        show_int_switchport_int = f_show_int_switchport[p]['interface']
        if f_switch_port == show_int_switchport_int:
            switchport_mode = f_show_int_switchport[p]['mode']
            return switchport_mode


def find_switchport_voice_vlan(f_show_int_switchport, f_switch_port):
    for p in range(len(f_show_int_switchport)):
        show_int_switchport_int = f_show_int_switchport[p]['interface']
        if f_switch_port == show_int_switchport_int:
            switchport_voice_vlan = f_show_int_switchport[p]['voice_vlan']
            return switchport_voice_vlan


def find_endpoint_mac_add(f_show_mac_add, f_switch_port):
    string_of_macs_per_port = ''
    for m in range(len(f_show_mac_add)):
        show_mac_port = f_show_mac_add[m]['destination_port']
        show_mac_mac = f_show_mac_add[m]['destination_address']
        show_mac_port_string = ''.join(show_mac_port)
        if f_switch_port == show_mac_port_string:
            string_of_macs_per_port = string_of_macs_per_port + show_mac_mac + '\n'
    new_string_of_macs_per_port = string_of_macs_per_port[:-1]
    return new_string_of_macs_per_port


def find_endpoint_mac_add_per_vlan(f_show_mac_add, f_switch_port, f_c):
    string_of_macs_per_port_per_vlan = ''
    for m in range(len(f_show_mac_add)):
        show_mac_port = f_show_mac_add[m]['destination_port']
        show_mac_mac = f_show_mac_add[m]['destination_address']
        show_mac_vlan = f_show_mac_add[m]['vlan']
        show_mac_port_string = ''.join(show_mac_port)
        if f_switch_port == show_mac_port_string:
            if f_c == show_mac_vlan:
                string_of_macs_per_port_per_vlan = string_of_macs_per_port_per_vlan + show_mac_mac + '\n'
    new_string_of_macs_per_port_per_vlan = string_of_macs_per_port_per_vlan[:-1]
    return new_string_of_macs_per_port_per_vlan

def mac_lookup(string_of_mac_per_port):
    list_of_mac_per_port = string_of_mac_per_port.splitlines()
    string_of_mac_lookup = ''
    for l in list_of_mac_per_port:
        try:
            string_of_mac_lookup = string_of_mac_lookup + MacLookup().lookup(l) + '\n'
        except:
            string_of_mac_lookup = string_of_mac_lookup + '!!!!!!' + '\n'
    new_string_of_mac_lookup = string_of_mac_lookup[:-1]
    return new_string_of_mac_lookup


k = 2

c93_port_pattern = re.compile('^(G|T)../(0)/')
i41_port_pattern = re.compile('^Gi1/(?:[1-9]|1[0-9]|2[0-4])$')
i40_port_pattern_uplink = re.compile('^(Gi1/[1-4]$)')
i_port_pattern_uplink = re.compile('^(Gi1/[1-2]$)')

# sheet_obj.cell(k, 6).alignment = Alignment(wrapText=True)

for x in switch_ips:
    switch_connection_parameters = {
        'ip': x,
        'device_type': 'cisco_ios',
        'username': '', # username
        'password': '' # password
    }

    ssh_session = ConnectHandler(**switch_connection_parameters)
    hostname = (ssh_session.find_prompt()).strip('>')
    print(f'======= CONNECTED TO {hostname} =======')
    show_inv = ssh_session.send_command('show inventory', use_textfsm=True)
    switch_model = show_inv[0]['pid']
    show_int_status = ssh_session.send_command('show interface status', use_textfsm=True)
    show_int_des = ssh_session.send_command('show interface description', use_textfsm=True)
    show_mac_add = ssh_session.send_command('show mac address-table', use_textfsm=True)
    show_int_switchport = ssh_session.send_command("show interface switchport", use_textfsm=True)

    if 'C9300' in switch_model:
        for y in range(len(show_int_status)):
            switch_port = show_int_status[y]['port']
            if re.match(c93_port_pattern, switch_port):
                port_vlan = show_int_status[y]['vlan']
                if port_vlan == 'trunk':
                    trunk_vlan_list = find_switch_port_trunk_vlans(show_int_switchport, switch_port)
                    trunk_vlan_list_expanded = expand_trunk_vlan_list(trunk_vlan_list)
                    for c in trunk_vlan_list_expanded:
                        sheet_obj.cell(k, 1).value = hostname
                        sheet_obj.cell(k, 2).value = switch_port
                        sheet_obj.cell(k, 3).value = find_switch_port_des(show_int_des, switch_port)
                        sheet_obj.cell(k, 4).value = show_int_status[y]['status']
                        sheet_obj.cell(k, 5).value = find_switchport_mode(show_int_switchport, switch_port)
                        sheet_obj.cell(k, 6).value = find_switchport_voice_vlan(show_int_switchport, switch_port)
                        sheet_obj.cell(k, 7).value = c
                        mac_string_per_vlan = find_endpoint_mac_add_per_vlan(show_mac_add, switch_port, c)
                        sheet_obj.cell(k, 8).value = mac_string_per_vlan
                        sheet_obj.cell(k, 9).value = mac_lookup(mac_string_per_vlan)
                        k = k + 1
                else:
                    sheet_obj.cell(k, 1).value = hostname
                    sheet_obj.cell(k, 2).value = switch_port
                    sheet_obj.cell(k, 3).value = find_switch_port_des(show_int_des, switch_port)
                    sheet_obj.cell(k, 4).value = show_int_status[y]['status']
                    sheet_obj.cell(k, 5).value = find_switchport_mode(show_int_switchport, switch_port)
                    sheet_obj.cell(k, 6).value = find_switchport_voice_vlan(show_int_switchport, switch_port)
                    sheet_obj.cell(k, 7).value = port_vlan
                    mac_string_per_vlan = find_endpoint_mac_add(show_mac_add, switch_port)
                    sheet_obj.cell(k, 8).value = mac_string_per_vlan
                    sheet_obj.cell(k, 9).value = mac_lookup(mac_string_per_vlan)
                    k = k + 1
    elif 'IE-4010' in switch_model:
        for y in range(len(show_int_status)):
            switch_port = show_int_status[y]['port']
            if re.match(i41_port_pattern, switch_port):
                port_vlan = show_int_status[y]['vlan']
                if port_vlan == 'trunk':
                    trunk_vlan_list = find_switch_port_trunk_vlans(show_int_switchport, switch_port)
                    trunk_vlan_list_expanded = expand_trunk_vlan_list(trunk_vlan_list)
                    for c in trunk_vlan_list_expanded:
                        sheet_obj.cell(k, 1).value = hostname
                        sheet_obj.cell(k, 2).value = switch_port
                        sheet_obj.cell(k, 3).value = find_switch_port_des(show_int_des, switch_port)
                        sheet_obj.cell(k, 4).value = show_int_status[y]['status']
                        sheet_obj.cell(k, 5).value = find_switchport_mode(show_int_switchport, switch_port)
                        sheet_obj.cell(k, 6).value = find_switchport_voice_vlan(show_int_switchport, switch_port)
                        sheet_obj.cell(k, 7).value = c
                        mac_string_per_vlan = find_endpoint_mac_add_per_vlan(show_mac_add, switch_port, c)
                        sheet_obj.cell(k, 8).value = mac_string_per_vlan
                        sheet_obj.cell(k, 9).value = mac_lookup(mac_string_per_vlan)
                        k = k + 1
                else:
                    sheet_obj.cell(k, 1).value = hostname
                    sheet_obj.cell(k, 2).value = switch_port
                    sheet_obj.cell(k, 3).value = find_switch_port_des(show_int_des, switch_port)
                    sheet_obj.cell(k, 4).value = show_int_status[y]['status']
                    sheet_obj.cell(k, 5).value = find_switchport_mode(show_int_switchport, switch_port)
                    sheet_obj.cell(k, 6).value = find_switchport_voice_vlan(show_int_switchport, switch_port)
                    sheet_obj.cell(k, 7).value = port_vlan
                    mac_string_per_vlan = find_endpoint_mac_add(show_mac_add, switch_port)
                    sheet_obj.cell(k, 8).value = mac_string_per_vlan
                    sheet_obj.cell(k, 9).value = mac_lookup(mac_string_per_vlan)
                    k = k + 1
    elif 'IE-4000' in switch_model:
        for y in range(len(show_int_status)):
            switch_port = show_int_status[y]['port']
            if not re.match(i40_port_pattern_uplink, switch_port):
                port_vlan = show_int_status[y]['vlan']
                if port_vlan == 'trunk':
                    trunk_vlan_list = find_switch_port_trunk_vlans(show_int_switchport, switch_port)
                    trunk_vlan_list_expanded = expand_trunk_vlan_list(trunk_vlan_list)
                    for c in trunk_vlan_list_expanded:
                        sheet_obj.cell(k, 1).value = hostname
                        sheet_obj.cell(k, 2).value = switch_port
                        sheet_obj.cell(k, 3).value = find_switch_port_des(show_int_des, switch_port)
                        sheet_obj.cell(k, 4).value = show_int_status[y]['status']
                        sheet_obj.cell(k, 5).value = find_switchport_mode(show_int_switchport, switch_port)
                        sheet_obj.cell(k, 6).value = find_switchport_voice_vlan(show_int_switchport, switch_port)
                        sheet_obj.cell(k, 7).value = c
                        mac_string_per_vlan = find_endpoint_mac_add_per_vlan(show_mac_add, switch_port, c)
                        sheet_obj.cell(k, 8).value = mac_string_per_vlan
                        sheet_obj.cell(k, 9).value = mac_lookup(mac_string_per_vlan)
                        k = k + 1
                else:
                    sheet_obj.cell(k, 1).value = hostname
                    sheet_obj.cell(k, 2).value = switch_port
                    sheet_obj.cell(k, 3).value = find_switch_port_des(show_int_des, switch_port)
                    sheet_obj.cell(k, 4).value = show_int_status[y]['status']
                    sheet_obj.cell(k, 5).value = find_switchport_mode(show_int_switchport, switch_port)
                    sheet_obj.cell(k, 6).value = find_switchport_voice_vlan(show_int_switchport, switch_port)
                    sheet_obj.cell(k, 7).value = port_vlan
                    mac_string_per_vlan = find_endpoint_mac_add(show_mac_add, switch_port)
                    sheet_obj.cell(k, 8).value = mac_string_per_vlan
                    sheet_obj.cell(k, 9).value = mac_lookup(mac_string_per_vlan)
                    k = k + 1
    elif 'IE' in switch_model:
        for y in range(len(show_int_status)):
            switch_port = show_int_status[y]['port']
            if not re.match(i_port_pattern_uplink, switch_port):
                port_vlan = show_int_status[y]['vlan']
                if port_vlan == 'trunk':
                    trunk_vlan_list = find_switch_port_trunk_vlans(show_int_switchport, switch_port)
                    trunk_vlan_list_expanded = expand_trunk_vlan_list(trunk_vlan_list)
                    for c in trunk_vlan_list_expanded:
                        sheet_obj.cell(k, 1).value = hostname
                        sheet_obj.cell(k, 2).value = switch_port
                        sheet_obj.cell(k, 3).value = find_switch_port_des(show_int_des, switch_port)
                        sheet_obj.cell(k, 4).value = show_int_status[y]['status']
                        sheet_obj.cell(k, 5).value = find_switchport_mode(show_int_switchport, switch_port)
                        sheet_obj.cell(k, 6).value = find_switchport_voice_vlan(show_int_switchport, switch_port)
                        sheet_obj.cell(k, 7).value = c
                        mac_string_per_vlan = find_endpoint_mac_add_per_vlan(show_mac_add, switch_port, c)
                        sheet_obj.cell(k, 8).value = mac_string_per_vlan
                        sheet_obj.cell(k, 9).value = mac_lookup(mac_string_per_vlan)
                        k = k + 1
                else:
                    sheet_obj.cell(k, 1).value = hostname
                    sheet_obj.cell(k, 2).value = switch_port
                    sheet_obj.cell(k, 3).value = find_switch_port_des(show_int_des, switch_port)
                    sheet_obj.cell(k, 4).value = show_int_status[y]['status']
                    sheet_obj.cell(k, 5).value = find_switchport_mode(show_int_switchport, switch_port)
                    sheet_obj.cell(k, 6).value = find_switchport_voice_vlan(show_int_switchport, switch_port)
                    sheet_obj.cell(k, 7).value = port_vlan
                    mac_string_per_vlan = find_endpoint_mac_add(show_mac_add, switch_port)
                    sheet_obj.cell(k, 8).value = mac_string_per_vlan
                    sheet_obj.cell(k, 9).value = mac_lookup(mac_string_per_vlan)
                    k = k + 1

    worksheet.save('scan.xlsx')
